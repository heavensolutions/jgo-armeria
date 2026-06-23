#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UNIFICADOR DE PRODUCTOS -- JGO Armeria -> Tienda Nube
=====================================================
Usage:
    python unificar_productos.py
"""

import csv
import re
import sys
import unicodedata
import json
import math
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

PROJECT = r"C:\Areas\02_Agencia\Clientes\JGO Armeria"

HEADERS = [
    "Identificador de URL", "Nombre", "Categorías",
    "Nombre de propiedad 1", "Valor de propiedad 1",
    "Nombre de propiedad 2", "Valor de propiedad 2",
    "Nombre de propiedad 3", "Valor de propiedad 3",
    "Precio", "Precio promocional", "Peso (kg)", "Alto (cm)", "Ancho (cm)", "Profundidad (cm)",
    "Stock", "SKU", "Código de barras", "Mostrar en tienda", "Envío sin cargo",
    "Descripción", "Tags", "Título para SEO", "Descripción para SEO",
    "Marca", "Producto Físico", "MPN (Número de pieza del fabricante)", "Sexo", "Rango de edad", "Costo"
]

# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────

def slugify(text):
    text = text.lower().strip()
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')

def parse_number(val, default=None):
    """Convert any numeric-ish value to float, return default on failure."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return default
        return float(val)
    s = str(val).strip().replace('$', '').replace('U$S', '').replace('U$', '').strip()
    s = s.replace(',', '.').strip()
    try:
        v = float(s)
        return v if v > 0 else default
    except (ValueError, TypeError):
        return default

def clean_price_str(val, min_p=10, max_p=100_000_000):
    """Validate and return price as clean string or empty."""
    v = parse_number(val)
    if v is None:
        return ''
    if v < min_p or v > max_p:
        return ''
    # Remove decimals for whole numbers
    if v == int(v):
        return str(int(v))
    return str(round(v, 2))

def normalize_marca(m):
    if not m:
        return 'Generica'
    m = m.strip().lower()
    mapping = {
        'glock': 'Glock', 'bersa': 'Bersa', 'taurus': 'Taurus',
        'beretta': 'Beretta', 'lube': 'Lube', 'olight': 'Olight',
        'earmor': 'Earmor', 'canik': 'Canik', 'f.a.b': 'FAB',
        'magpul': 'Magpul', 'dlg': 'DLG', 'key': 'KEY',
        'osight': 'Osight', 'b-optics': 'B-Optics',
        'snap caps': 'Snap Caps',
    }
    if m in mapping:
        return mapping[m]
    # Check partial match
    for k, v in mapping.items():
        if k in m:
            return v
    return m.title()

def normalize_categoria(sheet, descripcion, codigo):
    """Determine TN category from Bersa Shop sheet name and description."""
    s = sheet.strip().upper()
    desc = descripcion.lower()
    if s == 'OLIGHT':
        return 'Linternas'
    if s == 'EARMOR':
        return 'Proteccion Auditiva'
    if s == 'CUCHILLOS':
        return 'Cuchillos y Navajas'
    if s == 'SNAP CAPS':
        return 'Municiones y Snap Caps'
    if s == 'MERCHANDISING':
        return 'Merchandising'
    if s == 'OSIGHT  B-OPTICS':
        return 'Miras y Red Dots'
    if s in ('REPUESTOS', 'ACCESORIOS'):
        return 'Accesorios > Repuestos y Partes'
    if s == 'ACC. AR-BAR9':
        return 'Accesorios > AR-15 / BAR9'
    if s == 'CANIK - ACC':
        return 'Accesorios > Canik'
    if s == 'F.A.B':
        return 'Accesorios > Repuestos y Partes'
    return 'Accesorios'

# ─────────────────────────────────────────────────────────
# 1. CARGA DE PRODUCTOS BASE
# ─────────────────────────────────────────────────────────

def cargar_productos_base():
    productos = []
    try:
        with open(f'{PROJECT}/productos_jgo_armeria.csv', 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                productos.append(row)
        print(f"  -> Cargados {len(productos)} productos desde productos_jgo_armeria.csv")
    except FileNotFoundError:
        print("  ! productos_jgo_armeria.csv no encontrado")
    return productos

# ─────────────────────────────────────────────────────────
# 2. PRECIOS TAURUS (PDF)
# ─────────────────────────────────────────────────────────

def extraer_precios_taurus():
    import pdfplumber
    precios = {}
    pdf_path = f'{PROJECT}/listas/LISTA DE PRECIOS - TAURUS ABRIL 2026.pdf'
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                if 'EN STOCK' not in line and 'PROX. INGRESO' not in line:
                    continue
                # Find price: last numeric 3-5 digit value before status
                parts = line.split()
                price_candidates = []
                for i, p in enumerate(parts):
                    pc = p.replace(',', '.').replace('"', '')
                    try:
                        v = float(pc)
                        if 50 < v < 20000:
                            price_candidates.append((i, v))
                    except:
                        pass
                if not price_candidates:
                    continue
                price = price_candidates[-1][1]
                model_info = line[:line.rfind(parts[-1])].strip()

                # Identify model
                modelo = ''
                if 'G3C' in line or 'G3 C' in line or 'G3 Compacta' in line or 'G3 TORO' in line:
                    modelo = 'G3C'
                elif 'G3XL' in line or 'G3XL' in line:
                    modelo = 'G3XL'
                elif 'G2C' in line:
                    modelo = 'G2C'
                elif 'GX4' in line:
                    modelo = 'GX4'
                elif 'TS9' in line:
                    modelo = 'TS9'
                elif 'TH9' in line or 'THC' in line:
                    modelo = 'TH9'
                elif 'TH40' in line:
                    modelo = 'TH40'
                elif 'PT809' in line:
                    modelo = 'PT809E'
                elif 'PT1911' in line or 'COMMANDER' in line:
                    modelo = 'Commander 1911'
                elif 'OFFICER' in line:
                    modelo = 'Officer 1911'
                elif 'MTS9' in line:
                    modelo = 'MTS9'
                elif 'MTH40' in line:
                    modelo = 'MTH40C'
                elif 'T4' in line:
                    modelo = 'T4'
                elif 'T9' in line:
                    modelo = 'T9'
                elif 'M825' in line or '825' in line:
                    modelo = 'M825'
                elif '627' in line:
                    modelo = '627'
                elif '856' in line:
                    modelo = '856'
                elif 'M96' in line or '96' in line:
                    modelo = 'M96'
                elif 'M992' in line:
                    modelo = 'M992'
                elif 'M444' in line or '444' in line:
                    modelo = 'M444'
                elif 'M855' in line or '855' in line:
                    modelo = 'M855'
                elif 'M94' in line:
                    modelo = 'M94'
                elif 'M85' in line or '85S' in line:
                    modelo = 'M85'
                elif '605' in line:
                    modelo = '605'
                elif '689' in line:
                    modelo = '689'
                elif '669' in line:
                    modelo = '669'
                elif '44CP' in line or '44CP' in line:
                    modelo = '44CP'
                elif '82S' in line:
                    modelo = '82S'
                elif '970' in line:
                    modelo = '970'

                if modelo and modelo not in precios:
                    precios[modelo] = {'precio_usd': price, 'line': line[:100]}

    print(f"  -> Extraidos {len(precios)} modelos Taurus con precios del PDF")
    return precios

# ─────────────────────────────────────────────────────────
# 3. PRECIOS BERSA (PDF)
# ─────────────────────────────────────────────────────────

def extraer_precios_bersa():
    import pdfplumber
    productos = []
    pdf_path = f'{PROJECT}/listas/LISTA DE PRECIOS BERSA PISTOLAS Y ACCESORIOS - MI - JUNIO 2026.pdf'
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    current_calibre = ""
    current_modelo = ""
    seen_codes = set()

    for line in full_text.split('\n'):
        line = line.strip()
        if not line:
            continue

        cal_match = re.match(r'^CALIBRE\s+(.+)$', line, re.IGNORECASE)
        if cal_match:
            current_calibre = cal_match.group(1).strip()
            continue

        # Look for Bersa product codes: 0T32, 0T22, 0TPR9, etc.
        code_match = re.search(r'\b(0[A-Z0-9]+(?:[-/][A-Z0-9]+)?)\b', line)
        if not code_match:
            continue
        code = code_match.group(1)

        # Extract prices: two ARS values ($ XXX.XXX)
        prices = re.findall(r'\$\s*([\d.]+)', line)
        if len(prices) < 2:
            continue

        try:
            precio_sugerido = float(prices[1].replace('.', ''))
        except:
            continue

        # Skip if price is unrealistically low or high
        if precio_sugerido < 1000 or precio_sugerido > 10_000_000:
            continue

        if code in seen_codes:
            continue
        seen_codes.add(code)

        # Product name is everything before the code (approx)
        name = line[:line.find(code_match.group(0))].strip()
        # Clean up name (remove "PRECIO" and "SUGERIDO" artifacts)
        name = re.sub(r'^(PRECIO\s*)?(SUGERIDO\s*)?', '', name).strip()

        productos.append({
            'codigo': code,
            'nombre': name,
            'precio_ars': precio_sugerido,
            'calibre': current_calibre,
        })

    print(f"  -> Extraidos {len(productos)} productos Bersa del PDF")
    return productos

# ─────────────────────────────────────────────────────────
# 4. BERSA SHOP (XLSX) - Corregido para formulas
# ─────────────────────────────────────────────────────────

def extraer_bersa_shop():
    import openpyxl
    productos = []
    xlsx_path = f'{PROJECT}/listas/LISTA BERSA SHOP JUNIO 2026 - V1.xlsx'

    # Read the XLSX
    wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Column indices (0-based) for each sheet:
    # (code_col, name_col, features_col, [price_cols])
    SHEET_COLUMNS = {
        'REPUESTOS':     (1, 3, 2, [7, 6, 5]),       # code, desc, features, price(min_sug)
        'ACCESORIOS':    (1, 3, 2, [6, 5, 4]),       # code at col1, desc at col3, price at col6(min_sug)
        'CUCHILLOS':     (1, 3, 2, [5, 4]),          # code, desc at col3, price at col5(tienda)
        'EARMOR':        (1, 3, 6, [9, 8]),          # code, modelo at col3, features, price at col9(tienda)
        'OLIGHT':        (1, 3, 4, [7, 8, 6]),       # code, desc at col3, features, price at col7(tienda)
        'OSIGHT  B-OPTICS': (1, 4, 5, [8, 9, 7]),   # code, desc at col4, features, price at col8(tienda)
        'ACC. AR-BAR9':  (1, 4, 5, [7, 8, 6]),       # code at col1, desc/modelo at col4, features at col5, price at col7(tienda)
        'CANIK - ACC':   (1, 2, 5, [7, 8, 6]),       # code, modelo at col2, features at col5, price at col7(tienda)
        'F.A.B':         (1, 3, 5, [7, 8, 6]),       # code, desc at col3, features at col5, price at col7(tienda)
        'SNAP CAPS':     (1, 3, 2, [6, 5]),          # code, desc at col3, price at col6(tienda)
        'MERCHANDISING': (1, 4, 2, [7, 6, 5]),       # code, desc at col4, price at col7(min_sug)
    }

    for sheet_name in wb_data.sheetnames:
        ws_data = wb_data[sheet_name]
        cols = SHEET_COLUMNS.get(sheet_name)
        if not cols:
            continue

        code_col, name_col, feat_col, price_candidates = cols
        items = []
        header_row_found = False

        for row_data in ws_data.iter_rows(values_only=True):
            if not header_row_found:
                row_vals = [str(v or '').strip().lower() for v in row_data]
                joined = ' '.join(
                    unicodedata.normalize('NFKD', x).encode('ASCII', 'ignore').decode('ASCII')
                    for x in row_vals
                )
                if 'stock' in joined and any(h in joined for h in ['codigo', 'codico', 'sku']):
                    header_row_found = True
                continue

            if all(v is None or str(v).strip() == '' for v in row_data):
                continue

            def val(idx):
                return row_data[idx] if idx < len(row_data) else None
            def str_val(idx):
                return str(val(idx) or '').strip()

            codigo = str_val(code_col)
            nombre = str_val(name_col)
            features = str_val(feat_col)
            features_clean = re.sub(r'\s+', ' ', features).strip() if features else ''

            if not codigo or not nombre:
                continue

            descripcion = nombre
            if features_clean and len(features_clean) > 10:
                descripcion = f"{nombre} - {features_clean[:180]}"
            descripcion = re.sub(r'\s+', ' ', descripcion).strip()[:200]

            precio = ''
            for pc in price_candidates:
                pv = parse_number(val(pc))
                if pv:
                    p = clean_price_str(pv, min_p=200, max_p=100_000_000)
                    if p:
                        precio = p
                        break

            if not precio:
                continue

            # Determine brand
            marca = normalize_marca(sheet_name)
            if sheet_name == 'ACC. AR-BAR9':
                marca = 'Magpul'
            elif sheet_name == 'OLIGHT':
                marca = 'Olight'
            elif sheet_name == 'EARMOR':
                marca = 'Earmor'
            elif sheet_name == 'OSIGHT  B-OPTICS':
                marca = 'Osight'
            elif sheet_name == 'CANIK - ACC':
                marca = 'Canik'
            elif sheet_name == 'F.A.B':
                marca = 'FAB'
            elif sheet_name == 'SNAP CAPS':
                marca = 'Snap Caps'
            elif sheet_name in ('REPUESTOS', 'ACCESORIOS'):
                marca = 'Bersa'

            categoria = normalize_categoria(sheet_name, descripcion, codigo)

            if not precio:
                continue  # Skip products we can't price

            items.append({
                'codigo': codigo,
                'descripcion': re.sub(r'\s+', ' ', descripcion).strip()[:120],
                'precio': precio,
                'marca': marca,
                'categoria': categoria,
                'sheet': sheet_name,
            })

        if items:
            print(f"  -> {sheet_name}: {len(items)} productos")
            productos.extend(items)

    wb_data.close()
    print(f"  -> Total: {len(productos)} productos desde Bersa Shop XLSX")
    return productos

# ─────────────────────────────────────────────────────────
# 5. PRECIOS DE REFERENCIA
# ─────────────────────────────────────────────────────────

def get_reference_prices():
    """Return reference price tables by brand."""
    return {
        'Taurus': {  # USD - del PDF de Trompia
            'G3C': 440, 'G2C': 420, 'Commander 1911': 1250,
            'Officer 1911': 1000, 'MTS9': 450, 'MTH40C': 530,
            'M825': 850, '627': 830, '856': 850, 'M96': 850,
            'M992': 850, 'M444': 900, 'M855': 630, 'M94': 660,
            'M85': 630, '605': 670, '689': 900, '669': 950,
            '44CP': 900, '82S': 630, '970': 850,
        },
        'Glock': {  # USD - precios de referencia de mercado arg
            '17 gen3': 650, '17 gen5': 750, '19 gen5': 780,
            '19 gen4': 720, '19x arena': 850, '21 gen4': 800,
            '22 gen4': 750, '25': 600, '30 gen4': 850, '30s': 800,
            '35 gen4': 800, '43 gen4': 700, '44 gen4': 550,
            '45mos gen4': 850, '45compacta': 800, '48': 700, '48 dt': 720,
        },
        'Beretta': {  # USD - mercado arg
            'APX': 600, '92 Compact': 900, '92FS': 950,
            '9x9': 700, 'Storm': 750, 'APX Cany': 650,
            'APX Compact': 550, 'U22 NEOS': 500, '3032 Tomcat': 800,
            'APX AI': 600, '1301 Tactical': 1600,
            'A300 Ultim': 1200, 'A400 Lite': 1500, 'BRXI': 2000,
        },
        'Bersa': {  # ARS - precios sugeridos del PDF
            'bp9 fs': 970201, 'bp9cc': 740634, 'bp40cc': 735102,
            'bp380cc': 716553, 'tpr9': 880567, 'tpr9cx': 1011148,
            'tpr9cl': 1096276, 'tpr9c': 922953, 'tpr9x': 971875,
            'tpr9 fde': 941166, 'tpr9 odg': 941166,
            'tpr9 xt': 1514662, 'tpr40': 1023419, 'tpr40c': 1057436,
            'tpr45c': 1155777, 'thunder 32': 449167, 'thunder 22': 515646,
            'thunder 380': 490123, 'firestorm': 490122,
        }
    }

# ─────────────────────────────────────────────────────────
# 6. ASIGNAR PRECIOS A PRODUCTOS BASE
# ─────────────────────────────────────────────────────────

def assign_prices(productos, precios_taurus, precios_bersa):
    ref = get_reference_prices()

    # Build Bersa price map from reference prices (verified from PDF analysis)
    bersa_price_map = dict(ref['Bersa'])

    for p in productos:
        marca = normalize_marca(p.get('Marca', ''))
        nombre = p.get('Nombre', '')
        nombre_lower = nombre.lower()
        precio_actual = p.get('Precio', '')

        if precio_actual:
            continue

        precio_asignado = ''

        if marca == 'Taurus':
            for model_name, pr in ref['Taurus'].items():
                if model_name.lower() in nombre_lower:
                    precio_asignado = str(pr)
                    break
            if not precio_asignado and precios_taurus:
                for model, data in precios_taurus.items():
                    if model.lower() in nombre_lower:
                        precio_asignado = str(data['precio_usd'])
                        break

        elif marca == 'Glock':
            for model_name, pr in ref['Glock'].items():
                if model_name.lower() in nombre_lower:
                    precio_asignado = str(pr)
                    break

        elif marca == 'Beretta':
            for model_name, pr in ref['Beretta'].items():
                if model_name.lower() in nombre_lower:
                    precio_asignado = str(pr)
                    break

        elif marca == 'Bersa':
            # Try exact model matching first
            # Match Bersa prices - sort by keyword length (longest first) for specificity
            bersa_keywords = sorted(bersa_price_map.items(), key=lambda x: -len(x[0]))
            for kw, pr in bersa_keywords:
                if kw in nombre_lower:
                    precio_asignado = str(pr)
                    break

        p['Precio'] = precio_asignado
        p['Marca'] = marca

    return productos

# ─────────────────────────────────────────────────────────
# 7. GENERAR PRODUCTO TN
# ─────────────────────────────────────────────────────────

def make_tn_product(nombre, url, categoria, precio, marca, descripcion, tags='',
                    sku='', stock='5', mpn='', prop1='', val1='', prop2='', val2=''):
    """Create a Tienda Nube product dict."""
    seo_title = f"{nombre} | JGO Armeria"
    seo_desc = descripcion[:200] if descripcion else nombre
    return {
        "Identificador de URL": url,
        "Nombre": nombre[:150],
        "Categorías": categoria,
        "Nombre de propiedad 1": prop1, "Valor de propiedad 1": val1,
        "Nombre de propiedad 2": prop2, "Valor de propiedad 2": val2,
        "Nombre de propiedad 3": "", "Valor de propiedad 3": "",
        "Precio": precio, "Precio promocional": "",
        "Peso (kg)": "1", "Alto (cm)": "15", "Ancho (cm)": "10", "Profundidad (cm)": "30",
        "Stock": stock, "SKU": sku, "Código de barras": "",
        "Mostrar en tienda": "SI", "Envío sin cargo": "NO",
        "Descripción": descripcion, "Tags": tags or slugify(marca),
        "Título para SEO": seo_title,
        "Descripción para SEO": seo_desc,
        "Marca": marca, "Producto Físico": "SI",
        "MPN (Número de pieza del fabricante)": mpn or sku,
        "Sexo": "", "Rango de edad": "", "Costo": ""
    }

# ─────────────────────────────────────────────────────────
# 8. DEDUPLICACION
# ─────────────────────────────────────────────────────────

def deduplicar(productos):
    """Deduplicate by SKU > URL. Merge non-empty fields."""
    seen = {}
    dups = 0

    for p in productos:
        sku = (p.get('SKU') or '').strip()
        url = (p.get('Identificador de URL') or '').strip()
        key = f"sku:{sku}" if sku else f"url:{url}"

        if key in seen:
            existing = seen[key]
            # Merge: prefer non-empty, higher price
            for field in p:
                if field in ('Precio', 'Costo'):
                    try:
                        new_p = float(p.get(field) or 0)
                        old_p = float(existing.get(field, 0) or 0)
                        if new_p > old_p:
                            existing[field] = p[field]
                    except:
                        pass
                elif not existing.get(field) and p.get(field):
                    existing[field] = p[field]
            dups += 1
        else:
            seen[key] = p

    resultado = list(seen.values())
    print(f"  -> {dups} duplicados fusionados, {len(resultado)} productos unicos")
    return resultado

# ═════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("UNIFICADOR DE PRODUCTOS - JGO Armeria -> Tienda Nube")
    print("=" * 60)

    # 1. Cargar productos base
    print("\n[1/6] Cargar productos base")
    productos = cargar_productos_base()

    # 2. Precios Taurus
    print("\n[2/6] Extraer precios Taurus (PDF)")
    try:
        precios_taurus = extraer_precios_taurus()
    except Exception as e:
        print(f"  ! Error: {e}")
        precios_taurus = {}

    # 3. Precios Bersa
    print("\n[3/6] Extraer precios Bersa (PDF)")
    try:
        precios_bersa = extraer_precios_bersa()
    except Exception as e:
        print(f"  ! Error: {e}")
        precios_bersa = []

    # 4. Bersa Shop
    print("\n[4/6] Extraer productos Bersa Shop (XLSX)")
    try:
        bersa_shop = extraer_bersa_shop()
    except Exception as e:
        print(f"  ! Error: {e}")
        bersa_shop = []

    # 5. Asignar precios y construir catalogo final
    print("\n[5/6] Asignar precios y completar datos")

    productos_final = []

    # 5a. Procesar productos base con precios
    productos = assign_prices(productos, precios_taurus, precios_bersa)
    for p in productos:
        nombre = p.get('Nombre', '')
        marca = normalize_marca(p.get('Marca', ''))
        url = p.get('Identificador de URL', slugify(nombre))
        precio = p.get('Precio', '')
        sku = p.get('SKU', '')
        desc = p.get('Descripcion', nombre)
        cats = p.get('Categorías', '')
        tags = p.get('Tags', slugify(marca))

        # Fill empty fields
        cat = cats or ('Armas > Escopetas' if any(x in nombre.lower() for x in ['escopet', 'a300', 'a400', '1301', 'brxi'])
                       else 'Armas > Revolveres' if any(x in nombre.lower() for x in ['revolver', 'm8', 'm9', 'm4', 'm6', 'm3'])
                       else 'Armas > Pistolas')

        tp = make_tn_product(nombre, url, cat, precio, marca, desc, tags=tags, sku=sku)
        # Copy over property fields if they exist
        for prop in ['Nombre de propiedad 1', 'Valor de propiedad 1',
                     'Nombre de propiedad 2', 'Valor de propiedad 2',
                     'Nombre de propiedad 3', 'Valor de propiedad 3']:
            if p.get(prop):
                tp[prop] = p[prop]
        productos_final.append(tp)

    # 5b. Agregar productos de Bersa Shop
    shop_added = 0
    for item in bersa_shop:
        codigo = item['codigo']
        desc = item['descripcion']
        marca = item['marca']
        precio = item['precio']
        categoria = item['categoria']

        # Check if SKU already exists
        if any(p.get('SKU') == codigo for p in productos_final):
            continue
        # Check by URL
        url = slugify(f"{slugify(marca)}-{slugify(desc[:40])}")
        if any(p.get('Identificador de URL') == url for p in productos_final):
            continue

        nombre = f"{marca} {desc}" if not desc.startswith(marca) else desc
        if len(nombre) > 120:
            nombre = nombre[:117] + "..."

        tp = make_tn_product(nombre, url, categoria, precio, marca, desc,
                             tags=slugify(marca), sku=codigo, mpn=codigo)
        productos_final.append(tp)
        shop_added += 1

    print(f"  -> Agregados {shop_added} productos de Bersa Shop")

    # 6. Deduplicacion final
    print("\n[6/6] Deduplicacion final")
    productos_final = deduplicar(productos_final)

    # Sort
    productos_final.sort(key=lambda p: (p.get('Categorías', ''), p.get('Nombre', '')))

    # Write CSV
    output_path = f'{PROJECT}/productos_jgo_armeria_unificado.csv'
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS, delimiter=';',
                                quotechar='"', quoting=csv.QUOTE_ALL,
                                extrasaction='ignore')
        writer.writeheader()
        writer.writerows(productos_final)

    # Report
    total = len(productos_final)
    con_precio = sum(1 for p in productos_final if p.get('Precio'))
    sin_precio = [p for p in productos_final if not p.get('Precio')]
    cats = defaultdict(int)
    for p in productos_final:
        cats[p.get('Categorías', 'Sin categoría')] += 1

    print(f"\n{'=' * 60}")
    print(f"ARCHIVO GENERADO: {output_path}")
    print(f"Total productos: {total}")
    print(f"Con precio: {con_precio}/{total} ({con_precio*100//total}%)")

    if sin_precio:
        print(f"\nSin precio ({len(sin_precio)}):")
        for p in sin_precio[:10]:
            print(f"  - {p['Nombre']} (SKU: {p.get('SKU','')})")

    print(f"\nPor categoria:")
    for c, n in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")

    # Report JSON
    report = {
        'total_productos': total, 'con_precio': con_precio,
        'sin_precio': len(sin_precio),
        'por_categoria': dict(cats),
        'productos_sin_precio': [
            {'nombre': p['Nombre'], 'sku': p.get('SKU',''), 'marca': p.get('Marca','')}
            for p in sin_precio
        ]
    }
    with open(f'{PROJECT}/reporte_unificacion.json', 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\nReporte: {PROJECT}/reporte_unificacion.json")
    print(f"Listo! Subi 'productos_jgo_armeria_unificado.csv' a Tienda Nube.")

if __name__ == '__main__':
    main()
